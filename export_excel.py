import pandas as pd
import matplotlib.pyplot as plt
import os
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.drawing.image import Image

def create_chart(df, x, y, title, filename):
    plt.figure(figsize=(8, 5))
    if x not in df.columns or y not in df.columns:
        print(f"Colonnes '{x}' ou '{y}' non trouvées. Ignoré.")
        return
    df = df.sort_values(by=x)
    df.groupby(x)[y].mean().plot(kind="line", marker="o")
    plt.title(title)
    plt.ylabel(y)
    plt.xticks(rotation=45)
    plt.tight_layout()
    plt.savefig(filename)
    plt.close()

def create_threshold_plot(df, x, y, title, filename, seuils):
    if x not in df.columns or y not in df.columns:
        print(f"Données invalides pour {title}")
        return

    df[x] = pd.to_datetime(df[x])
    df = df.sort_values(by=x)
    df = df.groupby(x)[y].mean().reset_index()

    plt.figure(figsize=(10, 5))
    plt.plot(df[x], df[y], marker="o", label="Moyenne réelle", color="black")

    for val, label, color in seuils:
        plt.axhline(y=val, color=color, linestyle="--", label=label)

    plt.title(title)
    plt.xlabel("Date")
    plt.ylabel(y)
    plt.xticks(rotation=45)
    plt.legend()
    plt.tight_layout()
    plt.savefig(filename)
    plt.close()

def create_dashboard_with_alerts(wb):
    ws = wb.create_sheet("Dashboard")

    try:
        df_no2 = pd.read_csv("data_clean/NO2_alert_200_3days.csv")
        df_pm10 = pd.read_csv("data_clean/PM10_alerts.csv")
    except FileNotFoundError as e:
        ws["A1"] = f"Fichiers d'alerte non trouvés : {e.filename}"
        return

    no2_counts = df_no2["commune"].value_counts().reset_index()
    no2_counts.columns = ["commune", "nb_alertes_no2"]

    pm10_counts = df_pm10["commune"].value_counts().reset_index()
    pm10_counts.columns = ["commune", "nb_alertes_pm10"]

    summary = pd.merge(no2_counts, pm10_counts, on="commune", how="outer").fillna(0)

    for r in dataframe_to_rows(summary, index=False, header=True):
        ws.append(r)

    def make_plot(data, col_x, col_y, title, filename):
        if data.empty:
            print(f"Données vides pour : {title}")
            return
        plt.figure(figsize=(10, 5))
        plt.bar(data[col_x], data[col_y])
        plt.title(title)
        plt.xticks(rotation=45, ha="right")
        plt.tight_layout()
        plt.savefig(filename)
        plt.close()

    make_plot(no2_counts, "commune", "nb_alertes_no2", "Alertes NO2", "no2_alerts.png")
    make_plot(pm10_counts, "commune", "nb_alertes_pm10", "Alertes PM10", "pm10_alerts.png")

    if os.path.exists("no2_alerts.png"):
        ws.add_image(Image("no2_alerts.png"), "G2")
    if os.path.exists("pm10_alerts.png"):
        ws.add_image(Image("pm10_alerts.png"), "G30")

def add_normes_sheet(wb):
    ws = wb.create_sheet("Normes")
    normes_data = [
        ["Polluant", "Type de norme", "Valeur", "Unité", "Origine"],
        ["NO2", "Objectif de qualité (annuel)", "40", "μg/m³", "FR"],
        ["NO2", "Valeur limite horaire (18h/an)", "200", "μg/m³", "UE"],
        ["NO2", "Seuil d’information", "200", "μg/m³", "FR"],
        ["NO2", "Seuil d’alerte", "400", "μg/m³", "UE/FR"],
        ["PM10", "Objectif de qualité (annuel)", "30", "μg/m³", "FR"],
        ["PM10", "Valeur limite journalière (35j/an)", "50", "μg/m³", "UE"],
        ["PM10", "Seuil d’information", "50", "μg/m³", "FR"],
        ["PM10", "Seuil d’alerte", "80", "μg/m³", "FR"],
        ["O3", "Seuil d’alerte", "240", "μg/m³", "FR"],
        ["SO2", "Seuil d’alerte", "500", "μg/m³", "FR"],
    ]
    for row in normes_data:
        ws.append(row)
    ws.append([])
    ws.append(["Définitions :"])
    defs = [
        "Objectif de qualité : niveau souhaité à long terme",
        "Valeur limite : niveau à ne pas dépasser",
        "Seuil d’information : niveau avec risque pour groupes sensibles",
        "Seuil d’alerte : niveau nécessitant des mesures d’urgence"
    ]
    for defi in defs:
        ws.append([defi])

def run_export():
    data_dir = "data_clean"
    export_dir = os.path.join(os.path.dirname(__file__), "export_auto")
    os.makedirs(export_dir, exist_ok=True)
    output_file = os.path.join(export_dir, "export_analyse_air.xlsx")


    df_no2 = pd.read_csv(os.path.join(data_dir, "NO2_data.csv"))
    df_pm10 = pd.read_csv(os.path.join(data_dir, "PM10_data.csv"))
    df_pop = pd.read_csv(os.path.join(data_dir, "population_data.csv"))
    df_ent = pd.read_csv(os.path.join(data_dir, "enterprise_data.csv"))

    wb = Workbook()
    ws_no2 = wb.active
    ws_no2.title = "NO2"
    for r in dataframe_to_rows(df_no2.head(100), index=False, header=True):
        ws_no2.append(r)

    create_chart(df_no2, "date_local", "no2_value", "Évolution NO2", "no2_plot.png")
    create_threshold_plot(df_no2, "date_local", "no2_value", "NO2 avec seuils", "no2_seuils.png", [
        (40, "Valeur limite annuelle (40)", "blue"),
        (200, "Seuil info (200)", "orange"),
        (400, "Seuil alerte (400)", "red")
    ])
    if os.path.exists("no2_plot.png"):
        ws_no2.add_image(Image("no2_plot.png"), "H2")
    if os.path.exists("no2_seuils.png"):
        ws_no2.add_image(Image("no2_seuils.png"), "H20")

    ws_pm10 = wb.create_sheet("PM10")
    for r in dataframe_to_rows(df_pm10.head(100), index=False, header=True):
        ws_pm10.append(r)

    create_chart(df_pm10, "date_local", "pm10_value", "Évolution PM10", "pm10_plot.png")
    create_threshold_plot(df_pm10, "date_local", "pm10_value", "PM10 avec seuils", "pm10_seuils.png", [
        (50, "Seuil info (50)", "orange"),
        (80, "Seuil alerte (80)", "red")
    ])
    if os.path.exists("pm10_plot.png"):
        ws_pm10.add_image(Image("pm10_plot.png"), "H2")
    if os.path.exists("pm10_seuils.png"):
        ws_pm10.add_image(Image("pm10_seuils.png"), "H20")

    ws_pop = wb.create_sheet("Population")
    for r in dataframe_to_rows(df_pop, index=False, header=True):
        ws_pop.append(r)

    ws_ent = wb.create_sheet("Entreprises")
    top_activites = df_ent["activite_principale"].value_counts().head(10).reset_index()
    top_activites.columns = ["activite_principale", "count"]
    for r in dataframe_to_rows(top_activites, index=False, header=True):
        ws_ent.append(r)

    create_chart(top_activites, "activite_principale", "count", "Top activités entreprises", "ent_plot.png")
    if os.path.exists("ent_plot.png"):
        ws_ent.add_image(Image("ent_plot.png"), "E2")

    create_dashboard_with_alerts(wb)
    add_normes_sheet(wb)
    wb.save(output_file)
    print(f"Rapport Excel généré : {output_file}")

if __name__ == "__main__":
    run_export()
